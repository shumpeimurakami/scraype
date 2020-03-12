import requests
import xlrd
import xlwt
import json
import openpyxl
import re
from bs4 import BeautifulSoup as bs
import time


##ローカルにあるhandle.xlsxにアクセスする（一列目にハンドル名入力)
wb2 = openpyxl.load_workbook('/Users/days/Documents/google_search/handle.xlsx')
ws=wb2.get_sheet_by_name('Sheet1')
sheet2=wb2['Sheet1']

##excelの一列目のハンドルを取得し、三列目に出力する
for i in range(0,sheet2.max_row):
    time.sleep(1)
    print(sheet2.cell(row=i+1,column=1).value)    
    req = requests.get("https://www.google.com/search?q=" + sheet2.cell(row=i+1,column=1).value)
    req = req.text
    soup = bs(req,"html.parser")
    tags = soup.find_all("div", class_="ZINbbc xpd O9g5cc uUPGi")
    if(tags[0].find("div",class_="BNeawe vvjwJb AP7Wnd") != None):
        ##title = tags[0].find("div", class_="BNeawe vvjwJb AP7Wnd").string.split('(@')[0]
        title2 = tags[0].find("div", class_="BNeawe vvjwJb AP7Wnd").string 
        if "(@" in title2:
            title2 = title2.split('(@')[0]
        else:
            if "- Twitter" in title2:
                title2 = title2.split('-')[0]
            if "✓" in title2:
                title2 = title2.split('✓')[0]
        print(title2)
        sheet2.cell(row=i+1,column=3).value= title2
        wb2.save('/Users/days/Documents/google_search/handle.xlsx')
wb2.save('/Users/days/Documents/google_search/handle.xlsx')    
