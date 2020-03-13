import requests
import xlrd
import xlwt
import json
import openpyxl
import random
import time

url='https://twitter.com/'

req = requests.get(url,params={'q':'@ai_kayanon'})

req.url

print(req.url)

wb = xlrd.open_workbook('/Users/days/Documents/google_handle_name_search/handle.xlsx')
print(wb.sheet_names())
sheet= wb.sheet_by_name('Sheet1')
print(type(sheet))

book=xlwt.Workbook()
New_sheet = book.add_sheet('Sheet1')

wb2 = openpyxl.load_workbook('/Users/days/Documents/google_handle_name_search/handle.xlsx')
ws=wb2.get_sheet_by_name('Sheet1')
sheet2=wb2['Sheet1']

for i in range(0,sheet2.max_row):
    print(sheet2.cell(row=i+1,column=1).value)
    req = requests.get(url+sheet2.cell(row=i+1,column=1).value)
    print(req.url)
    sheet2.cell(row=i+1,column=3).value=req.url
    wb2.save('/Users/days/Documents/google_handle_name_search/handle.xlsx')
wb2.save('/Users/days/Documents/google_handle_name_search/handle.xlsx')    
